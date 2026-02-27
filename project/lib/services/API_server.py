from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi import HTTPException
from fastapi import Response, status
import threading
import json
import math
import paho.mqtt.client as mqtt
from uvicorn import run
import sqlite3
import time
from datetime import datetime, timedelta
import io
import csv
import redis
import requests
import pandas as pd
from typing import List
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from fpdf import FPDF
from fpdf.enums import XPos, YPos
import os # <--- เพิ่ม import os
from fastapi import APIRouter

class ExportRequest(BaseModel):
    start_time: str
    end_time: str
    step: str
    file_format: str
    variables: List[str]
    plant_name: str = "UTI Factory"
    units: List[str] = None

# 1. Config & Setup
MQTT_BROKER = "iicloud.tplinkdns.com"
MQTT_PORT = 7036
MQTT_USER = "mqtt_user"
MQTT_PASS = "ADMINktt5120@"

# ==============================================================================
# [FIXED] ตั้งค่า Path ของ Database ให้เป็นแบบตายตัว (Absolute Path)
# เพื่อป้องกันปัญหาข้อมูลหายเมื่อ Run จากต่าง Folder
# ==============================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__)) # หาตำแหน่งไฟล์ API_server.py
DB_NAME = os.path.join(BASE_DIR, "energy_data.db")    # บังคับสร้าง db ไว้ข้างๆ ไฟล์นี้เสมอ

print(f"--------------------------------------------------")
print(f"Database Path: {DB_NAME}") # แสดงตำแหน่งไฟล์ DB ให้เห็นชัดๆ
print(f"--------------------------------------------------")

# Redis Configuration
REDIS_HOST = "localhost"
REDIS_PORT = 6379
REDIS_DB = 0

WEATHER_API_KEY = '635c661512b0b802dcf857383d4a9ed4' 
WEATHER_CITIES = {
    "UTI": "Bang Sao Thong, Samut Prakan,TH",
    "TPI": "Bangkok,TH"
}

# เชื่อมต่อ Redis
try:
    redis_client = redis.Redis(host=REDIS_HOST, port=REDIS_PORT, db=REDIS_DB, decode_responses=True)
    redis_client.ping()
    print("\033[92m🗸\033[0m Connected to Redis")
except Exception as e:
    print(f"\033[91m𐄂\033[0m Failed to connect to Redis: {e}")

UTI_DEFAULT_KEYS = [
    # --- METER ---
    "METER_V1", "METER_V2", "METER_V3",
    "METER_I1", "METER_I2", "METER_I3",
    "METER_KW", "METER_TOTAL_KWH",
    "METER_EXPORT_KVARH", "METER_EXPORT_KWH", "METER_IMPORT_KVARH", "METER_IMPORT_KWH",
    "METER_TOTAL_KVARH", "METER_HZ", "METER_PF",
    "METER_I_TOTAL", "METER_KVAR", "METER_KW_INVERT", "METER_GRID_POWER_KW",

    # --- EMS (เบิ้ล EMS_EMS_) ---
    "EMS_AUTO_TARGET_SETPOINT_KW_ADD", # ตัวนี้ไม่เบิ้ลใน Redis
    "EMS_EMS_PV_TOTAL_ENERGY", "EMS_EMS_PV_DAILY_ENERGY", "EMS_EMS_LOAD_TOTAL_ENERGY", "EMS_EMS_LOAD_DAILY_ENERGY",
    "EMS_EMS_GRID_TOTAL_IMPORT_ENERGY", "EMS_EMS_GRID_DAILY_IMPORT_ENERGY", "EMS_EMS_GRID_TOTAL_EXPORT_ENERGY", "EMS_EMS_GRID_DAILY_EXPORT_ENERGY",
    "EMS_EMS_BESS_DAILY_CHARGE_ENERGY", "EMS_EMS_BESS_DAILY_DISCHARGE_ENERGY", "EMS_EMS_CO2_EQUIVALENT",
    "EMS_EMS_ENERGYPRODUCEDFROMPV_DAILY", "EMS_EMS_ENERGYFEEDTOGRID_DAILY", "EMS_EMS_ENERGYCONSUMPTION_DAILY",
    "EMS_EMS_ENERGYFEEDFROMGRID_DAILY", "EMS_EMS_SOLARPOWER_KW", "EMS_EMS_LOADPOWER_KW","EMS_EMS_BATTERYPOWER_KW",
    "EMS_EMS_ENERGYPRODUCEDFROMPV_KWH", "EMS_EMS_ENERGYFEEDFROMGRID_KWH", "EMS_EMS_ENERGYCONSUMPTION_KWH",
    "EMS_EMS_RENEWRATIODAILY","EMS_EMS_RENEWRATIOLIFETIME",

    # --- BESS (บางตัวเบิ้ล BESS_BESS_ บางตัวไม่เบิ้ล) ---
    "BESS_SOC", "BESS_SOH", "BESS_V", "BESS_I", "BESS_KW", "BESS_TEMPERATURE",
    "BESS_TOTAL_DISCHARGE", "BESS_TOTAL_CHARGE", "BESS_SOC_MAX", "BESS_SOC_MIN",
    "BESS_POWER_KW_INVERT", "BESS_MANUAL_POWER_SETPOINT", "BESS_PID_CYCLETIME",
    "BESS_PID_TD", "BESS_PID_TI", "BESS_PID_GAIN", "BESS_BESS_TEMP_AMBIENT",
    "BESS_BESS_ALARM", "BESS_BESS_FAULT", "BESS_BESS_COMMUNICATION_FAULT",

    # --- PV1-4 (เบิ้ล PV1_PV1_ ฯลฯ) ---
    "PV1_PV1_GRID_POWER_KW", "PV1_PV1_LOAD_POWER_KW", "PV1_PV1_DAILY_ENERGY_POWER_KWH", "PV1_PV1_TOTAL_ENERGY_POWER_KWH",
    "PV1_PV1_POWER_FACTOR", "PV1_PV1_REACTIVE_POWER_KVAR", "PV1_PV1_ACTIVE_POWER_KW", "PV1_PV1_FAULT", "PV1_PV1_COMMUNICATION_FAULT",
    
    "PV2_PV2_ENERGY_DAILY_KW", "PV2_PV2_LIFETIMEENERGYPRODUCTION_KWH_START", "PV2_PV2_LIFETIMEENERGYPRODUCTION_KWH",
    "PV2_PV2_REACTIVEPOWER_KW", "PV2_PV2_APPARENTPOWER_KW", "PV2_PV2_POWER_KW", "PV2_PV2_LIFETIMEENERGYPRODUCTION",
    "PV2_PV2_POWERFACTOR_PERCEN", "PV2_PV2_REACTIVEPOWER", "PV2_PV2_APPARENTPOWER", "PV2_PV2_POWER", "PV2_PV2_COMMUNICATION_FAULT",
    
    "PV3_PV3_TOTAL_POWER_YIELDS_REAL", "PV3_PV3_TOTAL_APPARENT_POWER_KW", "PV3_PV3_TOTAL_REACTIVE_POWER_KW", "PV3_PV3_TOTAL_ACTIVE_POWER_KW",
    "PV3_PV3_TOTAL_POWER_YIELDS", "PV3_PV3_DAILY_POWER_YIELDS", "PV3_PV3_NOMINAL_ACTIVE_POWER", "PV3_PV3_COMMUNICATION_FAULT",
    
    "PV4_PV4_TOTAL_POWER_YIELDS_REAL", "PV4_PV4_TOTAL_APPARENT_POWER_KW", "PV4_PV4_TOTAL_REACTIVE_POWER_KW", "PV4_PV4_TOTAL_ACTIVE_POWER_KW",
    "PV4_PV4_TOTAL_POWER_YIELDS", "PV4_PV4_DAILY_POWER_YIELDS", "PV4_PV4_NOMINAL_ACTIVE_POWER", "PV4_PV4_COMMUNICATION_FAULT",
    
    # --- WEATHER ---
    "WEATHER_Temp", "WEATHER_TempMin", "WEATHER_TempMax", "WEATHER_Humidity", "WEATHER_WindSpeed",
    "WEATHER_Sunrise", "WEATHER_Sunset", "WEATHER_FeelsLike", "WEATHER_Pressure", "WEATHER_Icon","WEATHER_City"
]

UTI_UNIT_MAPPING = {
    # --- METER ---
    "METER_V1": "V", "METER_V2": "V", "METER_V3": "V",
    "METER_I1": "A", "METER_I2": "A", "METER_I3": "A",
    "METER_KW": "kW", "METER_TOTAL_KWH": "kWh",
    "METER_EXPORT_KVARH": "kVarh", "METER_EXPORT_KWH": "kWh", 
    "METER_IMPORT_KVARH": "kVarh", "METER_IMPORT_KWH": "kWh",
    "METER_TOTAL_KVARH": "kVarh", "METER_HZ": "Hz", "METER_PF": "-",
    "METER_I_TOTAL": "A", "METER_KVAR": "kVar", "METER_KW_INVERT": "kW", "METER_GRID_POWER_KW": "kW",

    # --- EMS (เบิ้ล EMS_EMS_) ---
    "EMS_AUTO_TARGET_SETPOINT_KW_ADD": "kW",
    "EMS_EMS_PV_TOTAL_ENERGY": "kWh", "EMS_EMS_PV_DAILY_ENERGY": "kWh", "EMS_EMS_LOAD_TOTAL_ENERGY": "kWh", "EMS_EMS_LOAD_DAILY_ENERGY": "kWh",
    "EMS_EMS_GRID_TOTAL_IMPORT_ENERGY": "kWh", "EMS_EMS_GRID_DAILY_IMPORT_ENERGY": "kWh", "EMS_EMS_GRID_TOTAL_EXPORT_ENERGY": "kWh", "EMS_EMS_GRID_DAILY_EXPORT_ENERGY": "kWh",
    "EMS_EMS_BESS_DAILY_CHARGE_ENERGY": "kWh", "EMS_EMS_BESS_DAILY_DISCHARGE_ENERGY": "kWh", "EMS_EMS_CO2_EQUIVALENT": "kg",
    "EMS_EMS_ENERGYPRODUCEDFROMPV_DAILY": "kWh", "EMS_EMS_ENERGYFEEDTOGRID_DAILY": "kWh", "EMS_EMS_ENERGYCONSUMPTION_DAILY": "kWh",
    "EMS_EMS_ENERGYFEEDFROMGRID_DAILY": "kWh", "EMS_EMS_SOLARPOWER_KW": "kW", "EMS_EMS_LOADPOWER_KW": "kW", "EMS_EMS_BATTERYPOWER_KW": "kW",
    "EMS_EMS_ENERGYPRODUCEDFROMPV_KWH": "kWh", "EMS_EMS_ENERGYFEEDFROMGRID_KWH": "kWh", "EMS_EMS_ENERGYCONSUMPTION_KWH": "kWh",
    "EMS_EMS_RENEWRATIODAILY": "%", "EMS_EMS_RENEWRATIOLIFETIME": "%",

    # --- BESS (เบิ้ล BESS_BESS_ บางตัว) ---
    "BESS_SOC": "%", "BESS_SOH": "%", "BESS_V": "V", "BESS_I": "A", "BESS_KW": "kW", "BESS_TEMPERATURE": "°C",
    "BESS_TOTAL_DISCHARGE": "kWh", "BESS_TOTAL_CHARGE": "kWh", "BESS_SOC_MAX": "%", "BESS_SOC_MIN": "%",
    "BESS_POWER_KW_INVERT": "kW", "BESS_MANUAL_POWER_SETPOINT": "kW", "BESS_PID_CYCLETIME": "s",
    "BESS_PID_TD": "s", "BESS_PID_TI": "s", "BESS_PID_GAIN": "-", "BESS_BESS_TEMP_AMBIENT": "°C",
    "BESS_BESS_ALARM": "-", "BESS_BESS_FAULT": "-", "BESS_BESS_COMMUNICATION_FAULT": "-",

    # --- PV1 ---
    "PV1_PV1_GRID_POWER_KW": "kW", "PV1_PV1_LOAD_POWER_KW": "kW", "PV1_PV1_DAILY_ENERGY_POWER_KWH": "kWh", "PV1_PV1_TOTAL_ENERGY_POWER_KWH": "kWh",
    "PV1_PV1_POWER_FACTOR": "-", "PV1_PV1_REACTIVE_POWER_KVAR": "kVar", "PV1_PV1_ACTIVE_POWER_KW": "kW", 
    "PV1_PV1_FAULT": "-", "PV1_PV1_COMMUNICATION_FAULT": "-",

    # --- PV2 ---
    "PV2_PV2_ENERGY_DAILY_KW": "kWh", "PV2_PV2_LIFETIMEENERGYPRODUCTION_KWH_START": "kWh", "PV2_PV2_LIFETIMEENERGYPRODUCTION_KWH": "kWh",
    "PV2_PV2_REACTIVEPOWER_KW": "kVar", "PV2_PV2_APPARENTPOWER_KW": "kVA", "PV2_PV2_POWER_KW": "kW", "PV2_PV2_LIFETIMEENERGYPRODUCTION": "kWh",
    "PV2_PV2_POWERFACTOR_PERCEN": "%", "PV2_PV2_REACTIVEPOWER": "kVar", "PV2_PV2_APPARENTPOWER": "kVA", "PV2_PV2_POWER": "kW", "PV2_PV2_COMMUNICATION_FAULT": "-",

    # --- PV3 & PV4 ---
    "PV3_PV3_TOTAL_POWER_YIELDS_REAL": "kWh", "PV3_PV3_TOTAL_APPARENT_POWER_KW": "kVA", "PV3_PV3_TOTAL_REACTIVE_POWER_KW": "kVar", "PV3_PV3_TOTAL_ACTIVE_POWER_KW": "kW",
    "PV3_PV3_TOTAL_POWER_YIELDS": "kWh", "PV3_PV3_DAILY_POWER_YIELDS": "kWh", "PV3_PV3_NOMINAL_ACTIVE_POWER": "kW", "PV3_PV3_COMMUNICATION_FAULT": "-",
    
    "PV4_PV4_TOTAL_POWER_YIELDS_REAL": "kWh", "PV4_PV4_TOTAL_APPARENT_POWER_KW": "kVA", "PV4_PV4_TOTAL_REACTIVE_POWER_KW": "kVar", "PV4_PV4_TOTAL_ACTIVE_POWER_KW": "kW",
    "PV4_PV4_TOTAL_POWER_YIELDS": "kWh", "PV4_PV4_DAILY_POWER_YIELDS": "kWh", "PV4_PV4_NOMINAL_ACTIVE_POWER": "kW", "PV4_PV4_COMMUNICATION_FAULT": "-",

    # --- WEATHER ---
    "WEATHER_Temp": "°C", "WEATHER_TempMin": "°C", "WEATHER_TempMax": "°C", "WEATHER_Sunrise": "timestamp", "WEATHER_Sunset": "timestamp",
    "WEATHER_FeelsLike": "°C", "WEATHER_Humidity": "%", "WEATHER_Pressure": "hPa", "WEATHER_WindSpeed": "m/s",
    "WEATHER_Cloudiness": "%", "WEATHER_Icon": "-","WEATHER_City": "-"
}

# ==========================================
# 1. รายชื่อตัวแปรทั้งหมด (Keys) ของ TPI
# ==========================================
TPI_DEFAULT_KEYS = [
    # --- EMS (Energy Management System) ---
    "EMS_PLOAD",
    "EMS_KWHLOADTOTAL",
    "EMS_KWHLOADDAILY",
    "EMS_CO2E",
    "EMS_RENEWRATIO",
    "EMS_RENEWRATIOLIFETIME",

    # --- METER (Main Power Meter) ---
    "METER_P", "METER_Q", "METER_S", "METER_PF",
    "METER_KWHTOTAL", "METER_KWHPOS", "METER_KWHNEG",
    "METER_KWHTOTALDAILY", "METER_KWHPOSDAILY", "METER_KWHNEGDAILY",
    "METER_V1", "METER_V2", "METER_V3",
    "METER_V12", "METER_V23", "METER_V31",
    "METER_I1", "METER_I2", "METER_I3",

    # --- SOLAR (Inverters & EMI) ---
    "SOLAR_SOLAR1_EMI1_IRRADIANCETOTAL",
    "SOLAR_SOLAR1_EMI1_IRRADIANCEDAILY",
    "SOLAR_SOLAR1_EMI1_TEMPAMBIENT",
    "SOLAR_SOLAR1_EMI1_TEMPPV",
    "SOLAR_SOLAR1_LOGGER1_P",
    "SOLAR_SOLAR1_LOGGER1_Q",
    "SOLAR_SOLAR1_LOGGER1_PF",
    "SOLAR_SOLAR1_LOGGER1_KWHTOTAL",
    "SOLAR_SOLAR1_LOGGER1_KWHDAILY",
    "SOLAR_SOLAR1_LOGGER1_IDC",
    "SOLAR_SOLAR1_LOGGER1_V12",
    "SOLAR_SOLAR1_LOGGER1_V23",
    "SOLAR_SOLAR1_LOGGER1_V31",
    "SOLAR_SOLAR1_LOGGER1_I1",
    "SOLAR_SOLAR1_LOGGER1_I2",
    "SOLAR_SOLAR1_LOGGER1_I3",
    "SOLAR_SOLAR1_METER2_P", "SOLAR_SOLAR1_METER2_Q", "SOLAR_SOLAR1_METER2_S", "SOLAR_SOLAR1_METER2_PF",
    "SOLAR_SOLAR1_METER2_KWHTOTAL", "SOLAR_SOLAR1_METER2_KWHPOS", "SOLAR_SOLAR1_METER2_KWHNEG",
    "SOLAR_SOLAR1_METER2_V1", "SOLAR_SOLAR1_METER2_V2", "SOLAR_SOLAR1_METER2_V3",
    "SOLAR_SOLAR1_METER2_V12", "SOLAR_SOLAR1_METER2_V23", "SOLAR_SOLAR1_METER2_V31",
    "SOLAR_SOLAR1_METER2_I1", "SOLAR_SOLAR1_METER2_I2", "SOLAR_SOLAR1_METER2_I3",
    "SOLAR_SOLAR1_METER3_P", "SOLAR_SOLAR1_METER3_Q", "SOLAR_SOLAR1_METER3_S", "SOLAR_SOLAR1_METER3_PF",
    "SOLAR_SOLAR1_METER3_KWHTOTAL", "SOLAR_SOLAR1_METER3_KWHPOS", "SOLAR_SOLAR1_METER3_KWHNEG",
    "SOLAR_SOLAR1_METER3_V1", "SOLAR_SOLAR1_METER3_V2", "SOLAR_SOLAR1_METER3_V3",
    "SOLAR_SOLAR1_METER3_V12", "SOLAR_SOLAR1_METER3_V23", "SOLAR_SOLAR1_METER3_V31",
    "SOLAR_SOLAR1_METER3_I1", "SOLAR_SOLAR1_METER3_I2", "SOLAR_SOLAR1_METER3_I3",

    # --- BESS (Battery Energy Storage System) ---
    # SCU (System Control Unit)
    "BESS_SCU_P", "BESS_SCU_I", "BESS_SCU_SOC", "BESS_SCU_SOH", "BESS_SCU_PINVERT",
    "BESS_SCU_KWHCHARGETOTAL", "BESS_SCU_KWHDISCHARGETOTAL",
    "BESS_SCU_KWHCHARGEDAILY", "BESS_SCU_KWHDISCHARGEDAILY",
    
    # Racks (Individual Battery Racks 1-5)
    # หมายเหตุ: ใส่ตัวแปรตัวอย่างตามที่พบใน Log (สามารถก๊อปปี้เพิ่มให้ครบทุก Rack ได้)
    "BESS_RACK1_KWHCHARGETOTAL","BESS_RACK1_KWHDISCHARGETOTAL","BESS_RACK1_KWHCHARGEDAILY","BESS_RACK1_KWHDISCHARGEDAILY","BESS_RACK1_TIMECHARGE","BESS_RACK1_TIMEDISCHARGE",
    "BESS_RACK1_PCSCOMMFAULT","BESS_RACK1_PCSFAULT","BESS_RACK1_PCSALARM","BESS_RACK1_PCSDERATING","BESS_RACK1_PCSBOOTING","BESS_RACK1_PCSGRIDTIED",
    "BESS_RACK1_PCSOFFGRID","BESS_RACK1_PCSFAIL","BESS_RACK1_PCSONOFF","BESS_RACK1_PCSSTANDBY","BESS_RACK1_PCSCHARGING","BESS_RACK1_PCSDISCHARGING",
    "BESS_RACK1_PCSFULLYCHARGE","BESS_RACK1_PCSTOTALLYDISCHARGE",
    "BESS_RACK1_V", "BESS_RACK1_I", "BESS_RACK1_P", "BESS_RACK1_SOC", "BESS_RACK1_SOH", "BESS_RACK1_STATE",
    "BESS_RACK1_CELLV", "BESS_RACK1_CELLTEMP", 
        
    
    "BESS_RACK2_KWHCHARGETOTAL","BESS_RACK2_KWHDISCHARGETOTAL","BESS_RACK2_KWHCHARGEDAILY","BESS_RACK2_KWHDISCHARGEDAILY","BESS_RACK2_TIMECHARGE","BESS_RACK2_TIMEDISCHARGE",
    "BESS_RACK2_PCSCOMMFAULT","BESS_RACK2_PCSFAULT","BESS_RACK2_PCSALARM","BESS_RACK2_PCSDERATING","BESS_RACK2_PCSBOOTING","BESS_RACK2_PCSGRIDTIED",
    "BESS_RACK2_PCSOFFGRID","BESS_RACK2_PCSFAIL","BESS_RACK2_PCSONOFF","BESS_RACK2_PCSSTANDBY","BESS_RACK2_PCSCHARGING","BESS_RACK2_PCSDISCHARGING",
    "BESS_RACK2_PCSFULLYCHARGE","BESS_RACK2_PCSTOTALLYDISCHARGE",
    "BESS_RACK2_V", "BESS_RACK2_I", "BESS_RACK2_P", "BESS_RACK2_SOC", "BESS_RACK2_SOH", "BESS_RACK2_STATE",
    "BESS_RACK2_CELLV", "BESS_RACK2_CELLTEMP",

    "BESS_RACK3_KWHCHARGETOTAL","BESS_RACK3_KWHDISCHARGETOTAL","BESS_RACK3_KWHCHARGEDAILY","BESS_RACK3_KWHDISCHARGEDAILY","BESS_RACK3_TIMECHARGE","BESS_RACK3_TIMEDISCHARGE",
    "BESS_RACK3_PCSCOMMFAULT","BESS_RACK3_PCSFAULT","BESS_RACK3_PCSALARM","BESS_RACK3_PCSDERATING","BESS_RACK3_PCSBOOTING","BESS_RACK3_PCSGRIDTIED",
    "BESS_RACK3_PCSOFFGRID","BESS_RACK3_PCSFAIL","BESS_RACK3_PCSONOFF","BESS_RACK3_PCSSTANDBY","BESS_RACK3_PCSCHARGING","BESS_RACK3_PCSDISCHARGING",
    "BESS_RACK3_PCSFULLYCHARGE","BESS_RACK3_PCSTOTALLYDISCHARGE",
    "BESS_RACK3_V", "BESS_RACK3_I", "BESS_RACK3_P", "BESS_RACK3_SOC", "BESS_RACK3_SOH", "BESS_RACK3_STATE",
    "BESS_RACK3_CELLV", "BESS_RACK3_CELLTEMP",

    "BESS_RACK4_KWHCHARGETOTAL","BESS_RACK4_KWHDISCHARGETOTAL","BESS_RACK4_KWHCHARGEDAILY","BESS_RACK4_KWHDISCHARGEDAILY","BESS_RACK4_TIMECHARGE","BESS_RACK4_TIMEDISCHARGE",
    "BESS_RACK4_PCSCOMMFAULT","BESS_RACK4_PCSFAULT","BESS_RACK4_PCSALARM","BESS_RACK4_PCSDERATING","BESS_RACK4_PCSBOOTING","BESS_RACK4_PCSGRIDTIED",
    "BESS_RACK4_PCSOFFGRID","BESS_RACK4_PCSFAIL","BESS_RACK4_PCSONOFF","BESS_RACK4_PCSSTANDBY","BESS_RACK4_PCSCHARGING","BESS_RACK4_PCSDISCHARGING",
    "BESS_RACK4_PCSFULLYCHARGE","BESS_RACK4_PCSTOTALLYDISCHARGE",
    "BESS_RACK4_V", "BESS_RACK4_I", "BESS_RACK4_P", "BESS_RACK4_SOC", "BESS_RACK4_SOH", "BESS_RACK4_STATE",
    "BESS_RACK4_CELLV", "BESS_RACK4_CELLTEMP",

    "BESS_RACK5_KWHCHARGETOTAL","BESS_RACK5_KWHDISCHARGETOTAL","BESS_RACK5_KWHCHARGEDAILY","BESS_RACK5_KWHDISCHARGEDAILY","BESS_RACK5_TIMECHARGE","BESS_RACK5_TIMEDISCHARGE",
    "BESS_RACK5_PCSCOMMFAULT","BESS_RACK5_PCSFAULT","BESS_RACK5_PCSALARM","BESS_RACK5_PCSDERATING","BESS_RACK5_PCSBOOTING","BESS_RACK5_PCSGRIDTIED",
    "BESS_RACK5_PCSOFFGRID","BESS_RACK5_PCSFAIL","BESS_RACK5_PCSONOFF","BESS_RACK5_PCSSTANDBY","BESS_RACK5_PCSCHARGING","BESS_RACK5_PCSDISCHARGING",
    "BESS_RACK5_PCSFULLYCHARGE","BESS_RACK5_PCSTOTALLYDISCHARGE",
    "BESS_RACK5_V", "BESS_RACK5_I", "BESS_RACK5_P", "BESS_RACK5_SOC", "BESS_RACK5_SOH", "BESS_RACK5_STATE",
    "BESS_RACK5_CELLV", "BESS_RACK5_CELLTEMP",

    "WEATHER_Temp",
    "WEATHER_TempMin",
    "WEATHER_TempMax",
    "WEATHER_Sunrise",
    "WEATHER_Sunset",
    "WEATHER_FeelsLike",
    "WEATHER_Humidity",
    "WEATHER_Pressure",
    "WEATHER_WindSpeed",
    "WEATHER_Cloudiness",
    "WEATHER_Icon",
    "WEATHER_City"
]

ALL_PLANTS_KEYS = list(set(UTI_DEFAULT_KEYS + TPI_DEFAULT_KEYS))

# ==========================================
# 2. การจับคู่หน่วยของแต่ละตัวแปร (Unit Mapping) ของ TPI
# ==========================================
TPI_UNIT_MAPPING = {
    # --- EMS (Energy Management System) ---
    "EMS_PLOAD": "kW",
    "EMS_KWHLOADTOTAL": "kWh",
    "EMS_KWHLOADDAILY": "kWh",
    "EMS_CO2E": "kgCO2e",
    "EMS_RENEWRATIO": "%",
    "EMS_RENEWRATIOLIFETIME": "%",

    # --- METER (Main Power Meter) ---
    "METER_P": "kW",
    "METER_Q": "kVAR",
    "METER_S": "kVA",
    "METER_PF": "",
    "METER_KWHTOTAL": "kWh",
    "METER_KWHPOS": "kWh",
    "METER_KWHNEG": "kWh",
    "METER_KWHTOTALDAILY": "kWh",
    "METER_KWHPOSDAILY": "kWh",
    "METER_KWHNEGDAILY": "kWh",
    "METER_V1": "V", "METER_V2": "V", "METER_V3": "V",
    "METER_V12": "V", "METER_V23": "V", "METER_V31": "V",
    "METER_I1": "A", "METER_I2": "A", "METER_I3": "A",

    # --- SOLAR (Inverters & EMI) ---
    "SOLAR_SOLAR1_EMI1_IRRADIANCETOTAL": "W/m²",
    "SOLAR_SOLAR1_EMI1_IRRADIANCEDAILY": "kWh/m²",
    "SOLAR_SOLAR1_EMI1_TEMPAMBIENT": "°C",
    "SOLAR_SOLAR1_EMI1_TEMPPV": "°C",
    "SOLAR_SOLAR1_LOGGER1_P": "kW",
    "SOLAR_SOLAR1_LOGGER1_Q": "kVAR",
    "SOLAR_SOLAR1_LOGGER1_PF": "",
    "SOLAR_SOLAR1_LOGGER1_KWHTOTAL": "kWh",
    "SOLAR_SOLAR1_LOGGER1_KWHDAILY": "kWh",
    "SOLAR_SOLAR1_LOGGER1_IDC": "A",
    "SOLAR_SOLAR1_LOGGER1_V12": "V",
    "SOLAR_SOLAR1_LOGGER1_V23": "V",
    "SOLAR_SOLAR1_LOGGER1_V31": "V",
    "SOLAR_SOLAR1_LOGGER1_I1": "A",
    "SOLAR_SOLAR1_LOGGER1_I2": "A",
    "SOLAR_SOLAR1_LOGGER1_I3": "A",

    "WEATHER_Temp": "°C", "WEATHER_TempMin": "°C", "WEATHER_TempMax": "°C", "WEATHER_Sunrise": "timestamp", "WEATHER_Sunset": "timestamp",
    "WEATHER_FeelsLike": "°C", "WEATHER_Humidity": "%", "WEATHER_Pressure": "hPa", "WEATHER_WindSpeed": "m/s",
    "WEATHER_Cloudiness": "%", "WEATHER_Icon": "-","WEATHER_City": "-",

    # SOLAR Meter 2 & 3
    **{f"SOLAR_SOLAR1_METER{m}_{k}": v for m in [2, 3] for k, v in {
        "P": "kW", "Q": "kVAR", "S": "kVA", "PF": "",
        "KWHTOTAL": "kWh", "KWHPOS": "kWh", "KWHNEG": "kWh",
        "V1": "V", "V2": "V", "V3": "V", "V12": "V", "V23": "V", "V31": "V",
        "I1": "A", "I2": "A", "I3": "A"
    }.items()},

    # --- BESS (Battery Energy Storage System) ---
    # SCU
    "BESS_SCU_P": "kW",
    "BESS_SCU_I": "A",
    "BESS_SCU_SOC": "%",
    "BESS_SCU_SOH": "%",
    "BESS_SCU_PINVERT": "kW",
    "BESS_SCU_KWHCHARGETOTAL": "kWh",
    "BESS_SCU_KWHDISCHARGETOTAL": "kWh",
    "BESS_SCU_KWHCHARGEDAILY": "kWh",
    "BESS_SCU_KWHDISCHARGEDAILY": "kWh",

    # Racks 1-5 (ใช้ Loop เพื่อความกระชับและแม่นยำ)
    **{f"BESS_RACK{i}_{k}": v for i in range(1, 6) for k, v in {
        "KWHCHARGETOTAL": "kWh", "KWHDISCHARGETOTAL": "kWh",
        "KWHCHARGEDAILY": "kWh", "KWHDISCHARGEDAILY": "kWh",
        "TIMECHARGE": "min", "TIMEDISCHARGE": "min",
        "V": "V", "I": "A", "P": "kW", "SOC": "%", "SOH": "%",
        "CELLV": "V", "CELLTEMP": "°C",
        "STATE": "status",
        "PCSCOMMFAULT": "alarm", "PCSFAULT": "alarm", "PCSALARM": "alarm",
        "PCSDERATING": "status", "PCSBOOTING": "status", "PCSGRIDTIED": "status",
        "PCSOFFGRID": "status", "PCSFAIL": "alarm", "PCSONOFF": "status",
        "PCSSTANDBY": "status", "PCSCHARGING": "status", "PCSDISCHARGING": "status",
        "PCSFULLYCHARGE": "status", "PCSTOTALLYDISCHARGE": "status"
    }.items()}
}

print("Initializing Redis keys...")
pipe = redis_client.pipeline()
for plant in ["UTI", "TPI"]:
    # เลือก Keys ให้ตรงตามโรงงาน
    keys_to_init = UTI_DEFAULT_KEYS if plant == "UTI" else TPI_DEFAULT_KEYS
    for key in keys_to_init:
        pipe.setnx(f"{plant}:{key}", 0.0)
pipe.execute()
print("\033[92m🗸\033[0m Redis keys initialized complete.")

def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    clean_uti_keys = list(dict.fromkeys(UTI_DEFAULT_KEYS))
    clean_tpi_keys = list(dict.fromkeys(TPI_DEFAULT_KEYS))
    ALL_PLANTS_KEYS = list(dict.fromkeys(clean_uti_keys + clean_tpi_keys))
    
    col_defs = []
    seen_cols = set() # ตัวช่วยจำชื่อคอลัมน์ที่สร้างไปแล้ว (แบบพิมพ์เล็กทั้งหมด)
    
    for key in ALL_PLANTS_KEYS:
        key_lower = key.lower() # แปลงเป็นพิมพ์เล็กเพื่อเช็คตัวซ้ำ
        
        # ถ้ายังไม่เคยมีคอลัมน์นี้ ให้สร้างใหม่
        if key_lower not in seen_cols:
            seen_cols.add(key_lower)
            
            # กำหนด Data Type
            if key == "WEATHER_Icon" or "status" in key.lower() or "state" in key.lower() or "alarm" in key.lower():
                col_defs.append(f'"{key}" TEXT')
            else:
                col_defs.append(f'"{key}" REAL')
            
    columns_sql = ", ".join(col_defs)

    create_table_sql = f'''
        CREATE TABLE IF NOT EXISTS system_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp DATETIME,
            plant TEXT,
            {columns_sql}
        )
    '''
    
    try:
        cursor.execute(create_table_sql)
        conn.commit()
        print("\033[92m🗸\033[0m Database Initialized (Case-insensitive duplicates handled)")
    except sqlite3.OperationalError as e:
        print(f"\033[91m! Error during table creation:\033[0m {e}")
    finally:
        conn.close()

def init_db_wal_mode():
    max_retries = 5
    for i in range(max_retries):
        try:
            # เพิ่ม timeout=60 เพื่อให้โอกาสรอนานขึ้น
            with sqlite3.connect(DB_NAME, timeout=60) as conn:
                # สั่ง Commit เผื่อมี transaction ค้าง
                try: conn.execute("COMMIT") 
                except: pass
                
                cursor = conn.cursor()
                cursor.execute("PRAGMA journal_mode=WAL;")
                mode = cursor.fetchone()[0]
                
                if mode.upper() == 'WAL':
                    print(f"\033[92m🗸\033[0m Database WAL mode enabled. (Attempt {i+1})")
                    return
                else:
                    print(f"\033[93m⚠\033[0m WAL mode not set yet (Current: {mode}), retrying...")
                    
        except Exception as e:
            print(f"\033[93m⚠\033[0m Failed to enable WAL mode (Attempt {i+1}): {e}")
            time.sleep(1) # รอ 1 วินาทีก่อนลองใหม่
            
    print("\033[91m𐄂\033[0m Could not enable WAL mode after retries. System will continue but may be slow.")
    
init_db_wal_mode()

init_db()

app = FastAPI()
app.add_middleware(
    CORSMiddleware, 
    allow_origins=["*"], 
    allow_credentials=True, 
    allow_methods=["*"], 
    allow_headers=["*"],
    expose_headers=["Content-Disposition"]
)

def get_energy_at_time(cursor, target_datetime):
    # แปลง datetime เป็น string format ใน database
    target_str = target_datetime.strftime("%Y-%m-%d %H:%M:%S")
    
    # Query หาค่า EMS_EnergyProducedFromPV_kWh ที่เวลา <= target_time ที่ใกล้ที่สุด
    sql = """
        SELECT "EMS_EnergyProducedFromPV_kWh"
        FROM system_logs 
        WHERE timestamp <= ?
        ORDER BY timestamp DESC
        LIMIT 1
    """
    cursor.execute(sql, (target_str,))
    row = cursor.fetchone()
    if row and row[0] is not None:
        return float(row[0])
    return 0.0

# ==========================================
# 2. MQTT Logic (Write to Hot Data)
# ==========================================
def on_connect(client, userdata, flags, rc):
    print("Connected to MQTT Broker!")
    UTI_TOPICS = ["EMS/#", "BESS/#", "METER/#", "PV1/#", "PV2/#", "PV3/#", "PV4/#"]
    TPI_TOPICS = ["tpi/#"]
    for t in UTI_TOPICS + TPI_TOPICS: 
        client.subscribe(t)


def on_message(client, userdata, msg):
    try:
        original_topic = msg.topic 
        topic_lower = original_topic.lower()
        raw_payload = msg.payload.decode().strip()
        
        try:
            payload_value = float(raw_payload)
        except ValueError:
            return 

        if topic_lower.startswith("tpi/"):
            topic_core = original_topic[4:]
            key_name = topic_core.replace("/", "_").upper()
            redis_key = f"TPI:{key_name}"
            redis_client.set(redis_key, payload_value)
        else:
            if topic_lower.startswith("uti/"):
                topic_core = original_topic[4:]
            else:
                topic_core = original_topic
            key_name = topic_core.replace("/", "_").upper()
            redis_key = f"UTI:{key_name}"
            redis_client.set(redis_key, payload_value)

    except Exception as e:
        print(f"MQTT on_message Error: {e}")
# ==========================================
# Weather Fetcher Loop
# ==========================================
def weather_loop():
    print("\033[92m🗸\033[0m Weather Fetcher Started")
    while True:
        # วนลูปดึงข้อมูลทีละโรงงาน ตามเมืองที่ตั้งค่าไว้
        for plant, city in WEATHER_CITIES.items():
            try:
                url = f"https://api.openweathermap.org/data/2.5/weather?q={city}&units=metric&appid={WEATHER_API_KEY}"
                response = requests.get(url, timeout=10)
                
                if response.status_code == 200:
                    data = response.json()
                    weather_update = {
                        "WEATHER_City": city.replace(",TH", ""),
                        "WEATHER_Temp": data['main']['temp'],
                        "WEATHER_TempMin": data['main']['temp_min'],
                        "WEATHER_TempMax": data['main']['temp_max'],
                        "WEATHER_Sunrise": data['sys']['sunrise'],
                        "WEATHER_Sunset": data['sys']['sunset'],
                        "WEATHER_FeelsLike": data['main']['feels_like'],
                        "WEATHER_Humidity": data['main']['humidity'],
                        "WEATHER_Pressure": data['main']['pressure'],
                        "WEATHER_WindSpeed": data['wind']['speed'],
                        "WEATHER_Cloudiness": data.get('clouds', {}).get('all', 0),
                        "WEATHER_Icon": data['weather'][0]['icon']
                    }
                    pipe = redis_client.pipeline()
                    for k, v in weather_update.items():
                        redis_key = f"{plant}:{k}"
                        if k in ["WEATHER_Icon", "WEATHER_City"]:
                            pipe.set(redis_key, v)
                        else:
                            pipe.set(redis_key, round(float(v), 2))
                    pipe.execute()
                else:
                    print(f"Weather API Error ({plant} - {city}): {response.status_code}")
            
            except Exception as e:
                print(f"Error fetching weather for {plant}: {e}")
            
            # หน่วงเวลา 2 วินาทีระหว่างดึงแต่ละเมือง เพื่อป้องกัน API ของ OpenWeatherMap บล็อกรัว Request
            time.sleep(2)
        # เมื่อดึงครบ 2 โรงงานแล้ว ให้รอ 5 นาที (300 วินาที) แล้วค่อยทำใหม่
        time.sleep(300)

# สั่งรัน Weather Loop ใน Thread แยก
weather_thread = threading.Thread(target=weather_loop)
weather_thread.daemon = True
weather_thread.start()

# ==========================================
# 3. Background Tasks (Sync Hot -> Cold)
# ==========================================
# [EDITED] ฟังก์ชันนี้แก้ไขให้บันทึกทุก 5 นาที
def db_saver_loop():
    global last_mqtt_update
    print("\033[92m🗸\033[0m Database Saver Loop Started (Mode: Every 5 Minutes)")
    while True:
        try:
            now = datetime.now()
            # ตรวจสอบว่าถึงเวลาบันทึก (ทุก 5 นาที)
            if now.minute % 5 == 0:
                conn = sqlite3.connect(DB_NAME, timeout=30)
                cursor = conn.cursor()
                local_time_str = now.strftime("%Y-%m-%d %H:%M:%S")

                for current_plant in ["UTI", "TPI"]:
                    # --- [FIX 1] เลือก Keys ให้ตรงกับโรงงานที่กำลังบันทึก ---
                    keys_to_save = UTI_DEFAULT_KEYS if current_plant == "UTI" else TPI_DEFAULT_KEYS
                    
                    pipe = redis_client.pipeline()
                    for key in keys_to_save:
                        pipe.get(f"{current_plant}:{key}")
                    raw_values = pipe.execute()
                
                    vals = []
                    for idx, v in enumerate(raw_values):
                        key_name = keys_to_save[idx]
                        if key_name == "WEATHER_Icon":
                            vals.append(str(v) if v else "01d")
                        else:
                            try:
                                val_float = float(v) if v else 0.0
                                vals.append(round(val_float, 4))
                            except:
                                vals.append(0.0)

                    # --- [FIX 2] เพิ่มเครื่องหมาย ? และส่งค่า current_plant ให้ครบถ้วน ---
                    columns_str = ", ".join([f'"{k}"' for k in keys_to_save])
                    placeholders = ", ".join(["?" for _ in keys_to_save])
                    
                    # เพิ่ม ? ตัวที่สองสำหรับคอลัมน์ plant
                    sql = f'''
                        INSERT INTO system_logs (timestamp, plant, {columns_str})
                        VALUES (?, ?, {placeholders})
                    '''
                    
                    # ส่งค่า current_plant เข้าไปด้วย (ลำดับ: timestamp, plant, ...vals)
                    cursor.execute(sql, (local_time_str, current_plant, *vals))
                
                conn.commit()
                conn.close()
                print(f"\033[92m🗸\033[0m Archived data to DB at {local_time_str}")
                time.sleep(60) # ป้องกันการบันทึกซ้ำในนาทีเดียวกัน
            else:
                time.sleep(10)
        except Exception as e:
            print(f"Error syncing Hot-to-Cold data: {e}")
            time.sleep(10)

db_thread = threading.Thread(target=db_saver_loop)
db_thread.daemon = True
db_thread.start()

def start_mqtt():
    client = mqtt.Client()
    client.username_pw_set(MQTT_USER, MQTT_PASS)
    client.on_connect = on_connect
    client.on_message = on_message
    try: client.connect(MQTT_BROKER, MQTT_PORT, 60); client.loop_forever()
    except Exception as e: print(f"MQTT Error: {e}")

mqtt_thread = threading.Thread(target=start_mqtt)
mqtt_thread.daemon = True
mqtt_thread.start()

def on_disconnect(client, userdata, rc):
    if rc != 0:
        print("Unexpected disconnection. Attempting auto-reconnect...")
        try:
            client.reconnect()
        except:
            pass

mqtt_client = mqtt.Client()
mqtt_client.username_pw_set(MQTT_USER, MQTT_PASS)
mqtt_client.on_connect = on_connect
mqtt_client.on_message = on_message
mqtt_client.on_disconnect = on_disconnect

# ==========================================
# 4. API Endpoints
# ==========================================

@app.get("/api/dashboard")
def get_dashboard_data(plant: str = "UTI"):
    plant = plant.split('?')[0].upper() 
    KEYS_TO_USE = UTI_DEFAULT_KEYS
    if plant == "TPI":
        KEYS_TO_USE = TPI_DEFAULT_KEYS
        
    try:
        pipe = redis_client.pipeline()
        for k in KEYS_TO_USE:
            pipe.get(f"{plant}:{k}") 
        values = pipe.execute()
        data = {}
        success_count = 0 
        
        for i, key in enumerate(KEYS_TO_USE):
            val = values[i]
            if key in ["WEATHER_Icon", "WEATHER_City"]:
                if val:
                    data[key] = val.decode('utf-8') if isinstance(val, bytes) else str(val)
                else:
                    data[key] = "Unknown" if key == "WEATHER_City" else "01d"
            else:
                try:
                    if val is not None:
                        data[key] = round(float(val), 4)
                        success_count += 1
                    else:
                        data[key] = 0.0
                except:
                    data[key] = 0.0

        pv_daily = data.get("EMS_EnergyProducedFromPV_Daily", 0.0)
        load_daily = data.get("EMS_EnergyConsumption_Daily", 0.0)
        
        if load_daily > 0:
            data["EMS_RenewRatioDaily"] = round(pv_daily / load_daily, 4)
        else:
            data["EMS_RenewRatioDaily"] = 0.0
        pv_life = data.get("EMS_EnergyProducedFromPV_kWh", 0.0)
        load_life = data.get("EMS_EnergyConsumption_kWh", 0.0)
        
        if load_life > 0:
            data["EMS_RenewRatioLifetime"] = round(pv_life / load_life, 4)
        else:
            data["EMS_RenewRatioLifetime"] = 0.0
        return data
        
    except Exception as e:
        print(f"API Error: {e}")
        return {"error": str(e)}
    
# ==========================================
# 1. API หาช่วงวันที่ที่มีข้อมูล (Data Range)
# ==========================================
@app.get("/api/data_range")
def get_data_range():
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        # หาเวลาเริ่มต้นและสิ้นสุดที่มีข้อมูลใน DB
        cursor.execute("SELECT MIN(timestamp), MAX(timestamp) FROM system_logs")
        row = cursor.fetchone()
        conn.close()
        
        if row and row[0] and row[1]:
            return {"min_date": row[0], "max_date": row[1]}
        else:
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return {"min_date": now_str, "max_date": now_str}
    except Exception as e:
        return {"error": str(e)}
    
# ==========================================
# 2. API History ให้รับวันที่ (Daily)
# ==========================================
# เปลี่ยนชื่อจาก /api/history/today เป็น /api/history/daily
@app.get("/api/history/daily")
def get_daily_history(date: str = None):
    try:
        target_date = date if date else datetime.now().strftime("%Y-%m-%d")
        
        conn = sqlite3.connect(DB_NAME)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        sql = "SELECT * FROM system_logs WHERE date(timestamp) = ? ORDER BY timestamp ASC"
        cursor.execute(sql, (target_date,))
        rows = cursor.fetchall()
        conn.close()
        
        results = [dict(row) for row in rows]
        return results
    except Exception as e:
        return {"error": str(e)}
    
# ==========================================
# 3. API History (Monthly)
# ==========================================
@app.get("/api/history/monthly")
def get_month_history(year: int = None, month: int = None):
    try:
        now = datetime.now()
        target_year = year if year else now.year
        target_month = month if month else now.month
        target_str = f"{target_year}-{target_month:02d}"

        conn = sqlite3.connect(DB_NAME)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Query แบบ Group By วัน (เอาค่าล่าสุดของวัน)
        sql = """
        SELECT * FROM system_logs 
        WHERE id IN (
            SELECT MAX(id) 
            FROM system_logs 
            WHERE strftime('%Y-%m', timestamp) = ? 
            GROUP BY strftime('%d', timestamp)
        )
        ORDER BY timestamp ASC
        """
        cursor.execute(sql, (target_str,))
        rows = cursor.fetchall()
        conn.close()

        results = []
        for row in rows:
            d = dict(row)
            if "EMS_LoadPower_kW" in d and d["EMS_LoadPower_kW"] is not None:
                d["EMS_LoadPower_kW"] = abs(d["EMS_LoadPower_kW"])
            results.append(d)
        return results
    except Exception as e:
        return {"error": str(e)}

# ==========================================
# 4. API History (Yearly)
# ==========================================
@app.get("/api/history/yearly")
def get_year_history(year: int = None):
    try:
        now = datetime.now()
        target_year = year if year else now.year
        target_str = f"{target_year}"

        conn = sqlite3.connect(DB_NAME)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Query แบบ Group By เดือน (เอาค่าล่าสุดของเดือน)
        sql = """
        SELECT * FROM system_logs 
        WHERE id IN (
            SELECT MAX(id) 
            FROM system_logs 
            WHERE strftime('%Y', timestamp) = ? 
            GROUP BY strftime('%m', timestamp)
        )
        ORDER BY timestamp ASC
        """
        cursor.execute(sql, (target_str,))
        rows = cursor.fetchall()
        conn.close()

        results = []
        for row in rows:
            d = dict(row)
            if "EMS_LoadPower_kW" in d and d["EMS_LoadPower_kW"] is not None:
                d["EMS_LoadPower_kW"] = abs(d["EMS_LoadPower_kW"])
            results.append(d)
        return results
    except Exception as e:
        return {"error": str(e)}

# ==========================================
# 5. API สำหรับ Overview Chart (Daily/Monthly/Yearly)
# ==========================================
@app.get("/api/overview")
def get_overview_summary(mode: str = "daily", date_str: str = None):
    try:
        # -------------------------------------------------------
        # 1. โหมด Daily: ดึงค่า Realtime จาก Redis (เหมือนเดิม)
        # -------------------------------------------------------
        if mode == "daily":
            keys_map = [
                "PV_Daily_Energy",           
                "BESS_Daily_Charge_Energy",  
                "GRID_Daily_Export_Energy",  
                "Load_Daily_Energy",         
                "GRID_Daily_Import_Energy",  
                "BESS_Daily_Discharge_Energy"
            ]
            pipe = redis_client.pipeline()
            for k in keys_map: pipe.get(k)
            res = pipe.execute()
            data = [float(x) if x else 0.0 for x in res]
            return data

        # -------------------------------------------------------
        # 2. โหมด Monthly / Yearly: ดึงจาก SQLite
        # -------------------------------------------------------
        now = datetime.now()
        target_date = now 
        if date_str:
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d")
            except:
                pass # ถ้า format ผิด ให้ใช้เวลาปัจจุบัน

        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        # SQL Condition สำหรับกรองช่วงเวลา
        time_filter = ""
        debug_msg = ""

        if mode == "monthly":
            # กรอง "เดือน-ปี" เช่น '2026-01'
            t_str = target_date.strftime('%Y-%m')
            time_filter = f"strftime('%Y-%m', timestamp) = '{t_str}'"
            debug_msg = f"เดือน {t_str}"
        
        elif mode == "yearly":
            # กรอง "ปี" เช่น '2026'
            t_str = target_date.strftime('%Y')
            time_filter = f"strftime('%Y', timestamp) = '{t_str}'"
            debug_msg = f"ปี {t_str}"

        # -------------------------------------------------------
        # SQL LOGIC: 
        # 1. Subquery: หา MAX(id) ของแต่ละวัน (คือแถวสุดท้ายของวันนั้นๆ)
        # 2. Main Query: เอาค่าพลังงานของ id เหล่านั้นมารวมกัน (SUM)
        # -------------------------------------------------------
        sql = f"""
            SELECT 
                SUM("PV_Daily_Energy"),
                SUM("BESS_Daily_Charge_Energy"),
                SUM("GRID_Daily_Export_Energy"),
                SUM("Load_Daily_Energy"),
                SUM("GRID_Daily_Import_Energy"),
                SUM("BESS_Daily_Discharge_Energy")
            FROM system_logs 
            WHERE id IN (
                SELECT MAX(id) 
                FROM system_logs 
                WHERE {time_filter}
                GROUP BY strftime('%Y-%m-%d', timestamp)
            )
        """
        
        # --- เพิ่มส่วน Debug เพื่อเช็คว่าเจอวันไหนบ้าง ---
        check_sql = f"""
            SELECT strftime('%Y-%m-%d', timestamp), MAX(id) 
            FROM system_logs 
            WHERE {time_filter} 
            GROUP BY strftime('%Y-%m-%d', timestamp)
        """
        cursor.execute(check_sql)
        # ------------------------------------------------

        cursor.execute(sql)
        row = cursor.fetchone()
        conn.close()
        
        if row:
            # แปลง None เป็น 0.0
            result = [float(x) if x is not None else 0.0 for x in row]
            return result
        else:
            return [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]

    except Exception as e:
        print(f"Error overview: {e}")
        return [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]

@app.get("/api/export_csv")
def export_csv_data():
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM system_logs ORDER BY id DESC")
        rows = cursor.fetchall()
        if cursor.description is None: return {"error": "No data"}
        column_names = [description[0] for description in cursor.description]
        conn.close()

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(column_names)
        writer.writerows(rows)
        output.seek(0)
        
        filename = f"system_data_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/bill/reading_start")
def get_reading_start():
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        now = datetime.now()
        
        # =========================================================
        # Logic การหาวันเริ่มต้นรอบบิล (ตัดรอบทุกวันที่ 27)
        # =========================================================
        
        # กรณี A: วันนี้เป็นวันที่ 27 หรือมากกว่า (เช่น 28 ก.พ.)
        # รอบบิลเริ่มวันที่ 27 ของ "เดือนนี้"
        if now.day >= 27:
            start_date = datetime(now.year, now.month, 27, 0, 0, 0)
            
        # กรณี B: วันนี้ยังไม่ถึงวันที่ 27 (เช่น 15 ก.พ.)
        # รอบบิลเริ่มวันที่ 27 ของ "เดือนที่แล้ว"
        else:
            if now.month == 1:
                # ถ้าเป็นเดือนมกราคม ย้อนไปธันวาคมปีก่อนหน้า
                start_date = datetime(now.year - 1, 12, 27, 0, 0, 0)
            else:
                # เดือนปกติ ย้อนไปเดือนก่อนหน้า
                start_date = datetime(now.year, now.month - 1, 27, 0, 0, 0)
        
        # ใช้ฟังก์ชัน get_energy_at_time ที่มีอยู่แล้ว เพื่อดึงค่า ณ เวลานั้นๆ
        # ฟังก์ชันนี้จะหาค่าล่าสุดที่บันทึกไว้ ณ เวลา 00:00:00 หรือก่อนหน้านั้นที่ใกล้ที่สุด
        prev_read_val = get_energy_at_time(cursor, start_date)
        
        conn.close()

        return {"prev_read": prev_read_val}

    except Exception as e:
        print(f"Error fetching start reading: {e}")
        return {"prev_read": 0.0}
    
@app.get("/api/bill/calculate_tou")
def calculate_tou_units():
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        now = datetime.now()
        
        # 1. หาวันที่เริ่มต้นรอบบิล (วันที่ 27)
        if now.day >= 27:
            start_date = datetime(now.year, now.month, 27, 0, 0, 0)
        else:
            # ย้อนกลับไปเดือนก่อนหน้า
            if now.month == 1:
                start_date = datetime(now.year - 1, 12, 27, 0, 0, 0)
            else:
                start_date = datetime(now.year, now.month - 1, 27, 0, 0, 0)
        
        total_on_peak = 0.0
        total_off_peak = 0.0
        total_holiday = 0.0
        
        # 2. วนลูปตั้งแต่วันเริ่มต้น จนถึงวันนี้
        current_date = start_date
        # เราจะคำนวณทีละวัน (จบที่วันปัจจุบัน + 1 เพื่อให้ครอบคลุมวันนี้)
        end_date = datetime(now.year, now.month, now.day) + timedelta(days=1)
        
        while current_date < end_date:
            # current_date คือเวลา 00:00 ของวันนั้นๆ
            weekday = current_date.weekday() # 0=Mon, 1=Tue, ..., 5=Sat, 6=Sun
            
            # --- จันทร์ (0) ถึง ศุกร์ (4) ---
            if 0 <= weekday <= 4:
                # กำหนดเวลา 09:00 และ 22:00 ของวันนั้น
                time_00 = current_date.replace(hour=0, minute=0)
                time_09 = current_date.replace(hour=9, minute=0)
                time_22 = current_date.replace(hour=22, minute=0)
                
                # ถ้าเวลาที่จะดึง เป็นอนาคตเกินไป ให้ข้าม หรือใช้ค่าปัจจุบันแทน (ที่นี้ขอข้ามถ้าเกิน now)
                if time_00 <= now:
                     val_00 = get_energy_at_time(cursor, time_00)
                     
                     # 1. Off Peak (จ-ศ): 09:00 - 00:00
                     if time_09 <= now:
                         val_09 = get_energy_at_time(cursor, time_09)
                         # คำนวณ Off Peak
                         diff = val_09 - val_00
                         if diff > 0: total_off_peak += diff
                         
                         # 2. On Peak (จ-ศ): 22:00 - 09:00
                         if time_22 <= now:
                             val_22 = get_energy_at_time(cursor, time_22)
                         else:
                             # ถ้ายังไม่ถึง 22:00 ให้ใช้ค่าล่าสุด ณ ตอนนี้ (Realtime)
                             val_22 = get_energy_at_time(cursor, now)
                             
                         diff_on = val_22 - val_09
                         if diff_on > 0: total_on_peak += diff_on
                     else:
                         # กรณีวันนี้ยังไม่ถึง 09:00 (ได้ Off Peak บางส่วน)
                         val_now = get_energy_at_time(cursor, now)
                         diff = val_now - val_00
                         if diff > 0: total_off_peak += diff

            # --- เสาร์ (5) ---
            # Holiday คิดรวบยอด: จันทร์ถัดไป(00:00) - เสาร์(00:00)
            elif weekday == 5:
                time_sat_00 = current_date.replace(hour=0, minute=0)
                time_next_mon_00 = time_sat_00 + timedelta(days=2) # ข้ามอาทิตย์ไปจันทร์
                
                if time_sat_00 <= now:
                    val_sat = get_energy_at_time(cursor, time_sat_00)
                    
                    if time_next_mon_00 <= now:
                        val_mon = get_energy_at_time(cursor, time_next_mon_00)
                    else:
                        # ถ้ายังไม่ถึงเช้าวันจันทร์ ให้ใช้ค่าล่าสุด (Realtime)
                        val_mon = get_energy_at_time(cursor, now)
                    
                    diff_holiday = val_mon - val_sat
                    if diff_holiday > 0: total_holiday += diff_holiday
            
            # ขยับไปวันถัดไป
            current_date += timedelta(days=1)

        conn.close()
        
        return {
            "on_peak_unit": total_on_peak,
            "off_peak_unit": total_off_peak,
            "holiday_unit": total_holiday
        }

    except Exception as e:
        print(f"Error calculating TOU: {e}")
        return {"on_peak_unit": 0, "off_peak_unit": 0, "holiday_unit": 0}
    
@app.get("/api/data_range")
def get_data_range():
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        # หาเวลาน้อยสุดและมากสุด
        cursor.execute("SELECT MIN(timestamp), MAX(timestamp) FROM system_logs")
        result = cursor.fetchone()
        conn.close()

        min_date = result[0]
        max_date = result[1]

        # กรณีไม่มีข้อมูลใน DB เลย ให้ใช้เวลาปัจจุบันกัน Error
        if not min_date:
            min_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if not max_date:
            max_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        return {
            "min_date": min_date,
            "max_date": max_date
        }
    except Exception as e:
        print(f"Error getting data range: {e}")
        # Fallback กันตาย
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return {"min_date": now_str, "max_date": now_str}
    
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

@app.post("/api/export_custom")
def export_custom_data(req: ExportRequest, response: Response):
    try:
        print(f"Export Request: {req.start_time} to {req.end_time}, Step: {req.step}")

        # 1. Query ข้อมูล
        conn = sqlite3.connect(DB_NAME)
        cols = ", ".join(f'"{v}"' for v in req.variables) 
        query = f"""
            SELECT timestamp, {cols}
            FROM system_logs
            WHERE timestamp BETWEEN ? AND ?
            ORDER BY timestamp ASC
        """
        df = pd.read_sql_query(query, conn, params=(req.start_time, req.end_time))
        conn.close()

        if df.empty:
            response.status_code = status.HTTP_404_NOT_FOUND
            return {"detail": "No data found for the selected range"}

        # 2. Resample Data
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        df.set_index('timestamp', inplace=True)

        step_map = {
            '5 mins': '5min', '10 mins': '10min', '15 mins': '15min',
            '30 mins': '30min', '1 hour': '1h', '2 hours': '2h',
            '4 hours': '4h', '6 hours': '6h', '1 day': '1D'
        }
        pandas_step = step_map.get(req.step, '5min')
        df_resampled = df.resample(pandas_step).mean().fillna("Server Closed")

        new_col_names = [f"Point {i}" for i in range(1, len(df_resampled.columns) + 1)]
        df_resampled.columns = new_col_names

        # 3. สร้างไฟล์ Excel
        output = io.BytesIO()
        
        if req.file_format == 'Excel':
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                workbook = writer.book
                worksheet = workbook.create_sheet('ExportData')
                writer.sheets['ExportData'] = worksheet

                # --- Setup Styles ---
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                from openpyxl.utils import get_column_letter

                bold_font = Font(name='Arial', bold=True, size=8)
                center_align = Alignment(horizontal='center', vertical='center')
                left_align = Alignment(horizontal='left', vertical='center')
                right_align = Alignment(horizontal='right', vertical='center')
                
                normal_align = Alignment(horizontal='left', vertical='center', wrap_text=False)

                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                
                gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                blue_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")

                # --- ส่วนที่ 1: Header (Plant & Date) ---
                worksheet.row_dimensions[1].height = 40
                worksheet.merge_cells('A1:L1')
                worksheet.merge_cells('A2:B2')
                worksheet.merge_cells('C2:E2')
                worksheet.merge_cells('A3:B3')
                worksheet.merge_cells('C3:E3')
                cell_title = worksheet['A1']
                cell_title.value = req.plant_name
                cell_title.font = Font(name='Arial', bold=True, size=14)
                cell_title.alignment = center_align
                
                worksheet['A2'] = "Report Date :"
                worksheet['A2'].font = bold_font
                worksheet['A2'].alignment = right_align
                start_dt_obj = datetime.strptime(req.start_time, "%Y-%m-%d %H:%M:%S")
                end_dt_obj = datetime.strptime(req.end_time, "%Y-%m-%d %H:%M:%S")
                date_str = f"{start_dt_obj.strftime('%d %b %Y %H:%M')} - {end_dt_obj.strftime('%d %b %Y %H:%M')}"
                worksheet['C2'] = date_str
                worksheet['C2'].font = Font(name='Arial', bold=False, size=8)
                worksheet['C2'].alignment = left_align

                worksheet['A3'] = "Print Date :"
                worksheet['A3'].font = bold_font
                worksheet['A3'].alignment = right_align
                worksheet['C3'] = datetime.now().strftime('%d %b %Y %H:%M:%S')
                worksheet['C3'].font = Font(name='Arial', bold=False, size=8)
                worksheet['C3'].alignment = left_align

                # --- ส่วนที่ 2: Variable Table (Legend) ---
                start_meta_row = 5
                
                # 1. เขียน Header (แถวที่ 5)
                def write_header_row(start_col):
                    # Point
                    cp = worksheet.cell(row=start_meta_row, column=start_col, value="Point")
                    cp.font = bold_font; cp.border = thin_border; cp.alignment = center_align; cp.fill = gray_fill
                    # Name (Merge 4 cells: B-E หรือ H-K)
                    worksheet.merge_cells(start_row=start_meta_row, start_column=start_col+1, end_row=start_meta_row, end_column=start_col+4)
                    cn = worksheet.cell(row=start_meta_row, column=start_col+1, value="Name")
                    cn.font = bold_font; cn.alignment = center_align; cn.fill = gray_fill
                    for col in range(start_col+1, start_col+5):
                        worksheet.cell(row=start_meta_row, column=col).border = thin_border
                    # Unit
                    cu = worksheet.cell(row=start_meta_row, column=start_col+5, value="Unit")
                    cu.font = bold_font; cu.border = thin_border; cu.alignment = center_align; cu.fill = gray_fill

                write_header_row(1)  # ฝั่งซ้าย (A5-F5)
                write_header_row(7)  # ฝั่งขวา (G5-L5)

                # 2. เขียนข้อมูลหรือโครงสร้างว่าง (แถวที่ 6-10)
                for i in range(5): # วน 5 แถวเสมอ
                    current_r = start_meta_row + 1 + i
                    
                    # --- จัดการฝั่งซ้าย (A-F) ---
                    # ใส่เลข Point และตีกรอบเสมอ
                    c_p_l = worksheet.cell(row=current_r, column=1, value=i+1)
                    c_p_l.font = Font(name='Arial', size=8); c_p_l.border = thin_border; c_p_l.alignment = center_align; c_p_l.fill = gray_fill
                    
                    worksheet.merge_cells(start_row=current_r, start_column=2, end_row=current_r, end_column=5)
                    c_name_l = worksheet.cell(row=current_r, column=2)
                    c_unit_l = worksheet.cell(row=current_r, column=6)
                    
                    # ตีกรอบช่อง Name และ Unit เสมอ
                    for col in range(2, 6): worksheet.cell(row=current_r, column=col).border = thin_border
                    c_unit_l.border = thin_border

                    # ใส่ข้อมูลถ้ามีตัวแปรตัวที่ i
                    if i < len(req.variables):
                        var_name = req.variables[i]
                        c_name_l.value = var_name
                        
                        # ดึงหน่วยจาก UNIT_MAPPING ถ้าไม่มีให้ใส่ "-"
                        c_unit_l.value = UTI_UNIT_MAPPING.get(var_name, "-")
                        
                        c_name_l.font = Font(name='Arial', size=8); c_name_l.alignment = normal_align
                        c_unit_l.font = Font(name='Arial', size=8); c_unit_l.alignment = center_align

                    # --- จัดการฝั่งขวา (G-L) ---
                    # ใส่เลข Point 6-10 และตีกรอบเสมอ
                    c_p_r = worksheet.cell(row=current_r, column=7, value=i+6)
                    c_p_r.font = Font(name='Arial', size=8); c_p_r.border = thin_border; c_p_r.alignment = center_align; c_p_r.fill = gray_fill
                    
                    worksheet.merge_cells(start_row=current_r, start_column=8, end_row=current_r, end_column=11)
                    c_name_r = worksheet.cell(row=current_r, column=8)
                    c_unit_r = worksheet.cell(row=current_r, column=12)

                    # ตีกรอบช่อง Name และ Unit เสมอ
                    for col in range(8, 12): worksheet.cell(row=current_r, column=col).border = thin_border
                    c_unit_r.border = thin_border

                    # ใส่ข้อมูลถ้ามีตัวแปรตัวที่ i+5
                    idx_right = i + 5
                    if idx_right < len(req.variables):
                        var_name = req.variables[idx_right]
                        c_name_r.value = var_name
                        
                        # ดึงหน่วยจาก UNIT_MAPPING ถ้าไม่มีให้ใส่ "-"
                        c_unit_r.value = UTI_UNIT_MAPPING.get(var_name, "-")
                        
                        c_name_r.font = Font(name='Arial', size=8); c_name_r.alignment = normal_align
                        c_unit_r.font = Font(name='Arial', size=8); c_unit_r.alignment = center_align

                current_row = start_meta_row + 6

                # --- ส่วนที่ 3: Data Table ---
                data_start_row = current_row + 2
                df_resampled.columns = new_col_names
                df_resampled.index.name = None 
                data_start_row = current_row + 2 
                df_resampled.iloc[:, []].to_excel(writer, sheet_name='ExportData', startrow=data_start_row, startcol=0, header=False)
                df_resampled.to_excel(writer, sheet_name='ExportData', startrow=data_start_row, startcol=2, index=False, header=False)

                last_data_row = data_start_row + len(df_resampled)
                max_data_col = 12 

                # วนลูปจัดการ Format
                for r in range(data_start_row, last_data_row + 1):
                    
                    # แก้ไขจุดที่ 2: การ Merge และ Border (ต้องระบุพิกัดให้ชัดเจน)
                    worksheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                    
                    # ต้องดักจับเซลล์หลักหลัง Merge
                    cell_dt = worksheet.cell(row=r, column=1)
                    
                    # ตีกรอบทั้งช่องที่ 1 และ 2 (เพื่อให้เส้นรอบวง Merged Cell สมบูรณ์)
                    worksheet.cell(row=r, column=1).border = thin_border
                    worksheet.cell(row=r, column=2).border = thin_border
                    
                    if r == data_start_row:
                        # ส่วนหัวตาราง (แถวที่ 12)
                        cell_dt.value = "Date / Time"
                        cell_dt.font = bold_font
                        cell_dt.alignment = center_align
                        cell_dt.fill = blue_fill
                        worksheet.cell(row=r, column=2).fill = blue_fill # ใส่สีให้ครบช่องที่ merge
                    else:
                        # ส่วนข้อมูล (แถวที่ 13 เป็นต้นไป)
                        cell_dt.font = Font(name='Arial', size=8)
                        cell_dt.number_format = 'dd/mm/yyyy hh:mm'
                        cell_dt.alignment = center_align

                    # --- ส่วน Point 1-10 ---
                    for c in range(3, max_data_col + 1):
                        cell = worksheet.cell(row=r, column=c)
                        cell.border = thin_border
                        
                        if r == data_start_row:
                            cell.value = f"Point {c-2}"
                            cell.font = bold_font
                            cell.alignment = center_align
                            cell.fill = blue_fill 
                        else:
                            cell.font = Font(name='Arial', size=8)
                            cell.alignment = right_align
                            cell.number_format = '0.0000'

            output.seek(0)
            filename = f"{req.plant_name}-{req.start_time[:10]}-{req.step.replace(' ', '')}.xlsx"
            print(f"DEBUG: Generating filename -> {filename}")
            headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
            return StreamingResponse(
                iter([output.getvalue()]), 
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                headers=headers
            )
        
        elif req.file_format == 'PDF':
            # เปลี่ยน orientation เป็น 'P' (Portrait)
            pdf = FPDF(orientation='P', unit='mm', format='A4')
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            
            # --- Config Colors & Fonts ---
            pdf.set_font('helvetica', 'B', 14)
            # สีเทา (Legend)
            gray_color = (221, 221, 221)
            # สีฟ้า (Header)
            blue_color = (176, 196, 222)
            
            # --- 1. Title ---
            pdf.cell(0, 10, req.plant_name, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.ln(2)

            # --- 2. Info Block ---
            pdf.set_font('helvetica', 'B', 8)
            pdf.cell(30, 5, "Report Date :", align='R')
            pdf.set_font('helvetica', '', 8)
            
            start_dt_obj = datetime.strptime(req.start_time, "%Y-%m-%d %H:%M:%S")
            end_dt_obj = datetime.strptime(req.end_time, "%Y-%m-%d %H:%M:%S")
            date_str = f"{start_dt_obj.strftime('%d %b %Y %H:%M')} - {end_dt_obj.strftime('%d %b %Y %H:%M')}"
            pdf.cell(60, 5, date_str, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            pdf.set_font('helvetica', 'B', 8)
            pdf.cell(30, 5, "Print Date :", align='R')
            pdf.set_font('helvetica', '', 8)
            pdf.cell(60, 5, datetime.now().strftime('%d %b %Y %H:%M:%S'), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.ln(5)

            # --- 3. Legend Table ---
            # ปรับความกว้างให้พอดีกับแนวตั้ง (Usable Width ~190mm)
            # แบ่งซ้ายขวา: Side Width = 10+45+15 = 70mm
            # 2 ข้าง = 140mm + Gap 10mm = 150mm (เหลือที่ว่างสบายๆ)
            col_w_pt = 15
            col_w_nm = 64
            col_w_un = 15
            gap = 0 
            
            # Header Row for Legend
            pdf.set_fill_color(*gray_color)
            pdf.set_font('helvetica', 'B', 8)
            
            # Left Header
            pdf.cell(col_w_pt, 6, "Point", border=1, align='C', fill=True)
            pdf.cell(col_w_nm, 6, "Name", border=1, align='C', fill=True)
            pdf.cell(col_w_un, 6, "Unit", border=1, align='C', fill=True)
            
            #pdf.cell(gap, 6, "", border=0) # Gap
            
            # Right Header
            pdf.cell(col_w_pt, 6, "Point", border=1, align='C', fill=True)
            pdf.cell(col_w_nm, 6, "Name", border=1, align='C', fill=True)
            pdf.cell(col_w_un, 6, "Unit", border=1, align='C', fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            # Rows (Loop 5 times)
            for i in range(5):
                pdf.set_font('helvetica', '', 8)
                
                # --- Left Side ---
                idx_left = i
                name_l = req.variables[idx_left] if idx_left < len(req.variables) else ""
                unit_l = UTI_UNIT_MAPPING.get(name_l, "-") if name_l else ""
                
                pdf.set_fill_color(*gray_color)
                pdf.cell(col_w_pt, 6, str(i+1), border=1, align='C', fill=True)
                pdf.cell(col_w_nm, 6, name_l, border=1, align='L')
                pdf.cell(col_w_un, 6, unit_l, border=1, align='C')
                
                #pdf.cell(gap, 6, "", border=0)

                # --- Right Side ---
                idx_right = i + 5
                name_r = req.variables[idx_right] if idx_right < len(req.variables) else ""
                unit_r = UTI_UNIT_MAPPING.get(name_r, "-") if name_r else ""
                
                pdf.set_fill_color(*gray_color)
                pdf.cell(col_w_pt, 6, str(i+6), border=1, align='C', fill=True)
                pdf.cell(col_w_nm, 6, name_r, border=1, align='L')
                pdf.cell(col_w_un, 6, unit_r, border=1, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            pdf.ln(5)

            # --- 4. Data Table ---
            # ปรับความกว้างสำหรับแนวตั้ง:
            # Date = 30mm
            # Values = 16mm * 10 columns = 160mm
            # Total = 190mm (พอดีหน้ากระดาษเป๊ะ)
            w_date = 30
            w_val = 16 
            
            # Header
            pdf.set_fill_color(*blue_color)
            pdf.set_font('helvetica', 'B', 7) # ลด font header เล็กน้อย
            
            pdf.cell(w_date, 8, "Date / Time", border=1, align='C', fill=True)
            for i in range(10):
                pdf.cell(w_val, 8, f"Point {i+1}", border=1, align='C', fill=True)
            pdf.ln()
            
            # Data Rows
            pdf.set_font('helvetica', '', 7) # Font เนื้อหาขนาด 7
            
            if 'timestamp' not in df_resampled.columns:
                df_resampled.reset_index(inplace=True)

            for _, row in df_resampled.iterrows():
                # Date
                ts = row['timestamp']
                date_str = ts.strftime('%d/%m/%Y %H:%M') if hasattr(ts, 'strftime') else str(ts)
                pdf.cell(w_date, 6, date_str, border=1, align='C')
                
                # Values (10 Columns)
                for i in range(10):
                    if i < len(new_col_names):
                        val = row[new_col_names[i]]
                        val_str = f"{val:.4f}" if isinstance(val, (int, float)) else str(val)
                    else:
                        val_str = ""
                    
                    pdf.cell(w_val, 6, val_str, border=1, align='R')
                
                pdf.ln()

            # Output PDF
            pdf_output = io.BytesIO()
            pdf_bytes = pdf.output()
            pdf_output.write(pdf_bytes)
            pdf_output.seek(0)
            
            filename = f"{req.plant_name}-{req.start_time[:10]}.pdf"
            headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
            return StreamingResponse(pdf_output, media_type='application/pdf', headers=headers)

    except Exception as e:
        print(f"Export Error: {e}")
        response.status_code = status.HTTP_500_INTERNAL_SERVER_ERROR
        return {"detail": str(e)}
    
@app.get("/api/check_db_tables")
def check_db_tables():
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        # 1. ดูรายชื่อตารางทั้งหมด
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        
        db_structure = {}
        
        for table in tables:
            table_name = table[0]
            
            # 2. ดูชื่อคอลัมน์ในแต่ละตาราง
            cursor.execute(f"PRAGMA table_info({table_name})")
            columns = cursor.fetchall()
            column_names = [col[1] for col in columns]
            
            db_structure[table_name] = column_names
            
        conn.close()
        return {"status": "ok", "tables": db_structure}

    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/holidays/{year}")
def get_holidays(year: str):
    target_url = f"https://gateway.api.bot.or.th/financial-institutions-holidays/?year={year}"
    token = "eyJvcmciOiI2NzM1NzgwZWM4YzFlYjAwMDEyYTM3NzEiLCJpZCI6IjNhNGViOGU0YTY5NjQ5ZmJhMDU3MjlmMThiZmRiOTQzIiwiaCI6Im11cm11cjEyOCJ9"
    
    current_headers = {
        'X-IBM-Client-Id': token,
        'Authorization': f'Bearer {token}',
        'accept': 'application/json'
    }

    try:
        resp = requests.get(target_url, headers=current_headers, timeout=10)
        
        if resp.status_code == 200:
            res_data = resp.json()
            h_list = res_data.get('result', {}).get('data', [])
            
            h_dates = []
            h_details = {}
            
            for d in h_list:
                date_str = d.get('Date')
                if date_str and date_str != f"{year}-01-02":
                    h_dates.append(date_str)
                    h_details[date_str] = d.get('HolidayDescriptionThai', 'วันหยุด')
            return {
                "status": "ok",
                "holidays": h_dates,
                "holiday_details": h_details}
        else:
            return {"status": "error", "message": f"BOT API Error {resp.status_code}"}
            
    except Exception as e:
        return {"status": "error", "message": str(e)}

if __name__ == "__main__":
    print("Initializing Database...")
    init_db_wal_mode()
    print("Starting Server...")
    run(app, host="0.0.0.0", port=8000)